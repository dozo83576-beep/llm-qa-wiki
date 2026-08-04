"""
build-qa-report — собирает XLSX-отчёт формата Test Management из JSON.

Рантайм-нейтрально: только Python 3 + openpyxl, без библиотек конкретного агента.
Спецификация формата — docs/14-templates/xlsx-test-management.md.

Состав книги: Data (скрытый справочник) + Test Cases + Test Run + Bug Reports.
Листа Evidence нет и картинки не встраиваются — доказательства идут ссылками
в колонке Attachments/Links.

Использование:
    python tools/build-qa-report.py --input _qa-report.json --out QA_Проект_2026-07-30.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Не установлен openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

# ---------------------------------------------------------------- справочники

LOOKUPS = [
    ("ResultOptions", ["Pass", "Fail", "Blocked", "N/A"]),
    ("SeverityOptions", ["Critical", "Major", "Minor", "Trivial"]),
    ("PriorityOptions", ["High", "Medium", "Low"]),
    ("StatusOptions", ["Open", "In Progress", "Fixed", "Closed", "Rejected", "Duplicate"]),
]

RESULT_VALUES = LOOKUPS[0][1]
SEVERITY_VALUES = LOOKUPS[1][1]
PRIORITY_VALUES = LOOKUPS[2][1]

# Заказы ведёт один человек. Имя в поле Tester — его, а не безличное «QA-инженер»:
# отчёт подписан тем, кто отвечает за результат. Значение из JSON перекрывает умолчание.
DEFAULT_TESTER = "Николай"

# Двуязычные шапки: английский технический термин + русский перевод второй строкой.
# Ровно этот вид согласован с заказчиком, менять формулировки нельзя.
CASE_HEADERS = [
    ("Case ID\n(Номер кейса)", 12),
    ("Title\n(Название кейса)", 34),
    ("Preconditions\n(Предусловия)", 38),
    ("Steps\n(Шаги)", 52),
    ("Expected Result\n(Ожидаемый результат)", 46),
    ("Postconditions\n(Постусловия)", 38),
    ("Priority\n(Приоритет)", 12),
]

RUN_HEADERS = [
    ("Session ID\n(Номер сессии)", 12),
    ("Date\n(Время)", 20),
    ("Session Name\n(Имя сессии)", 40),
    ("Tester\n(Тестировщик)", 18),
    ("Case ID\n(Номер кейса)", 12),
    ("Result\n(Результат)", 13),
    ("Comment\n(Комментарий)", 62),
    ("Bug ID\n(Номер бага)", 12),
]

BUG_HEADERS = [
    ("Bug ID\n(Номер бага)", 12),
    ("Title\n(Название бага)", 38),
    ("Related Case ID\n(Связанные кейсы)", 30),
    ("Session ID\n(Номер сессии)", 12),
    ("Preconditions\n(Предусловия)", 40),
    ("Error Step\n(Шаг с ошибкой)", 40),
    ("Expected Result\n(Ожидаемый результат)", 44),
    ("Actual Result\n(Фактический результат)", 52),
    ("Severity\n(Серьёзность)", 13),
    ("Priority\n(Приоритет)", 12),
    ("Environment\n(Окружение)", 42),
    ("Attachments/Links\n(Скриншоты/Ссылки)", 38),
    ("Date Reported\n(Дата обнаружения)", 20),
]

# ------------------------------------------------------------------- стили

HEADER_FILL = PatternFill("solid", fgColor="DBE5F1")
HEADER_FONT = Font(bold=True, size=11, color="000000")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(*[Side(style="thin", color="A6A6A6")] * 4)

BODY_FONT = Font(size=10, color="000000")
# Короткие поля — идентификаторы, статусы, даты — стоят по центру: колонка узкая, значение
# читается как метка. Длинные тексты выключены влево и прижаты к нижнему краю: абзац в пять
# строк по центру читать тяжело, глазу не за что зацепиться в начале строки.
BODY_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN_TEXT = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

# Номера столбцов, которые выравниваются по центру. Остальные — текстовые.
CENTERED_CASE_COLS = {1, 7}                       # Case ID, Priority
CENTERED_RUN_COLS = {1, 2, 4, 5, 6, 8}            # Session ID, Date, Tester, Case ID, Result, Bug ID
CENTERED_BUG_COLS = {1, 3, 4, 9, 10, 11, 12, 13}  # Bug ID, Related Case ID, Session ID, Severity,
                                                  # Priority, Environment, Attachments/Links, Date Reported
BODY_BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)

LINK_FONT = Font(size=10, color="0563C1", underline="single")

# Даты в книге читает заказчик, а не машина: привычный ему день.месяц.год.
# На вход по-прежнему принимается ISO — разбором занимается parse_date.
DATE_FORMAT = "dd.mm.yyyy"

RESULT_STYLES = {
    "Pass": ("C6EFCE", "006100"),
    "Fail": ("FFC7CE", "9C0006"),
    "Blocked": ("FFEB9C", "9C6500"),
    "N/A": ("D9E1F2", "44546A"),
}


class ReportError(Exception):
    """Ошибка входных данных — сообщение показывается пользователю как есть."""


# --------------------------------------------------------------- вход/разбор

def parse_date(value, field: str) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text:
        raise ReportError(f"Не заполнена дата в поле '{field}'.")
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ReportError(f"Дата '{text}' в поле '{field}' не разбирается. Ожидается YYYY-MM-DD или YYYY-MM-DD HH:MM.")


def join_steps(value, field: str) -> str:
    """Шаги живут в одной ячейке нумерованным списком через перенос строки."""
    if isinstance(value, list):
        parts = [str(item).strip() for item in value if str(item).strip()]
        if not parts:
            raise ReportError(f"Пустой список шагов в поле '{field}'.")
        return "\n".join(parts)
    text = str(value or "").strip()
    if not text:
        raise ReportError(f"Не заполнено поле '{field}'.")
    return text


def require(mapping: dict, key: str, where: str) -> str:
    value = str(mapping.get(key) or "").strip()
    if not value:
        raise ReportError(f"{where}: не заполнено обязательное поле '{key}'.")
    return value


def validate_input(data: dict) -> None:
    """Проверки, которые дешевле поймать до сборки книги, чем после."""
    cases = data.get("cases") or []
    run = data.get("run") or []
    defects = data.get("defects") or []

    if not cases:
        raise ReportError("В отчёте нет ни одного тест-кейса.")
    if not run:
        raise ReportError("В отчёте нет ни одной строки прогона.")

    for index, case in enumerate(cases, start=1):
        expected = f"TC-{index:03d}"
        actual = require(case, "id", f"Кейс #{index}")
        if actual != expected:
            raise ReportError(
                f"Нарушена сплошная нумерация кейсов: кейс #{index} имеет id '{actual}', ожидается '{expected}'."
            )
        priority = require(case, "priority", f"Кейс {actual}")
        if priority not in PRIORITY_VALUES:
            raise ReportError(f"Кейс {actual}: недопустимый Priority '{priority}'. Допустимо: {', '.join(PRIORITY_VALUES)}.")

    case_ids = [str(case["id"]).strip() for case in cases]
    defect_ids = {require(defect, "id", f"Дефект #{i}") for i, defect in enumerate(defects, start=1)}

    if len(run) != len(cases):
        raise ReportError(f"Прогон покрывает {len(run)} кейсов из {len(cases)} — должны быть покрыты все.")

    for index, row in enumerate(run, start=1):
        case_id = require(row, "case", f"Строка прогона #{index}")
        if case_id != case_ids[index - 1]:
            raise ReportError(
                f"Строка прогона #{index} ссылается на '{case_id}', а по порядку ожидается '{case_ids[index - 1]}'. "
                "Test Run должен покрывать все кейсы в том же порядке."
            )
        result = require(row, "result", f"Строка прогона {case_id}")
        if result not in RESULT_VALUES:
            raise ReportError(f"Строка прогона {case_id}: недопустимый Result '{result}'. Допустимо: {', '.join(RESULT_VALUES)}.")

        bug = str(row.get("bug") or "").strip()
        if result == "Fail":
            if not bug:
                raise ReportError(f"Строка прогона {case_id}: Fail без Bug ID.")
            if bug not in defect_ids:
                raise ReportError(f"Строка прогона {case_id}: Bug ID '{bug}' не найден среди дефектов.")
        elif result == "Pass" and bug:
            raise ReportError(f"Строка прогона {case_id}: Pass ошибочно связан с багом '{bug}'.")

    for index, defect in enumerate(defects, start=1):
        defect_id = str(defect["id"]).strip()
        severity = require(defect, "severity", f"Дефект {defect_id}")
        if severity not in SEVERITY_VALUES:
            raise ReportError(f"Дефект {defect_id}: недопустимый Severity '{severity}'. Допустимо: {', '.join(SEVERITY_VALUES)}.")
        priority = require(defect, "priority", f"Дефект {defect_id}")
        if priority not in PRIORITY_VALUES:
            raise ReportError(f"Дефект {defect_id}: недопустимый Priority '{priority}'. Допустимо: {', '.join(PRIORITY_VALUES)}.")
        rows = defect.get("rows") or []
        if not rows:
            raise ReportError(f"Дефект {defect_id}: не задано ни одной строки (rows). Один дефект раскладывается построчно по затронутым элементам.")
        for row in rows:
            related = str(row.get("case") or "").strip()
            if related and related not in case_ids:
                raise ReportError(f"Дефект {defect_id}: ссылка на несуществующий кейс '{related}'.")


# ---------------------------------------------------------------- построение

def style_header(ws, headers, height=46) -> None:
    for col, (text, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = height
    ws.freeze_panes = "A2"


# Ширина столбца в символах и высота строки в пунктах связаны через размер шрифта тела:
# в 10 pt строка занимает ~13.2 pt по высоте, а в одну «единицу ширины» помещается ~1 символ.
LINE_HEIGHT_PT = 13.2
# Отступы держим минимальными: лишнего воздуха в ячейке быть не должно, но текст не должен
# упираться в границу.
ROW_PADDING_PT = 2.0
COL_PADDING = 1
MIN_COL_WIDTH = 9
# Без потолка столбец с шагами кейса растянулся бы на пол-экрана и книгу пришлось бы
# листать вбок. Текст при этом не теряется: он переносится, а высота строки считается ниже.
MAX_COL_WIDTH = 58


def wrapped_lines(value, width: int) -> int:
    """Сколько строк займёт значение в ячейке заданной ширины.

    Ширина столбца считается в условных символах, а шрифт пропорциональный: строка из «Ш» и «ы»
    шире строки из «i» и «л». Берём 0.9 от номинала — ошибка уходит в сторону лишнего сантиметра
    высоты, а не обрезанного текста.
    """
    text = "" if value is None else str(value)
    per_line = max(1, int(width * 0.9))
    total = 0
    for segment in text.splitlines() or [""]:
        total += max(1, -(-len(segment) // per_line))
    return max(1, total)


def fit_columns(ws, headers, first_row: int, last_row: int) -> list[int]:
    """Ширина столбца — по самой длинной строке его содержимого, в пределах потолка.

    Заголовок учитывается тоже: подпись столбца не должна обрезаться. Ширина «ровно под
    текст» экономит место там, где значения короткие (Severity, Priority, даты), и не даёт
    книге разъехаться там, где они длинные.
    """
    widths = []
    for col in range(1, len(headers) + 1):
        longest = max((len(part) for part in str(headers[col - 1][0]).splitlines()), default=0)
        for row in range(first_row, last_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            longest = max(longest, max((len(part) for part in str(value).splitlines()), default=0))
        width = min(MAX_COL_WIDTH, max(MIN_COL_WIDTH, longest + COL_PADDING))
        ws.column_dimensions[get_column_letter(col)].width = width
        widths.append(width)
    return widths


def fit_rows(ws, first_row: int, last_row: int, widths: list[int]) -> None:
    """Высота строки — ровно под самую высокую ячейку, чтобы текст был виден целиком."""
    for row in range(first_row, last_row + 1):
        lines = max(wrapped_lines(ws.cell(row=row, column=col).value, widths[col - 1])
                    for col in range(1, len(widths) + 1))
        ws.row_dimensions[row].height = round(lines * LINE_HEIGHT_PT + ROW_PADDING_PT, 1)


def style_body(ws, first_row: int, last_row: int, col_count: int, centered: set[int]) -> None:
    for row in range(first_row, last_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            # Ячейкам со ссылкой шрифт не трогаем: BODY_FONT затёр бы синий
            # подчёркнутый вид, и ссылка перестала бы читаться как ссылка.
            if not cell.hyperlink:
                cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN_CENTER if col in centered else BODY_ALIGN_TEXT
            cell.border = BODY_BORDER


def add_dropdown(ws, lookup: str, values: list[str], cell_range: str) -> None:
    """Выпадающий список по именованному диапазону.

    showErrorMessage обязателен: без него Excel показывает список, но молча
    принимает любой посторонний ввод — и отчёт уезжает заказчику с «Passed»
    вместо «Pass», а условное форматирование по нему не срабатывает.
    """
    dv = DataValidation(
        type="list",
        formula1=f"={lookup}",
        allow_blank=False,
        showDropDown=False,  # инвертировано в формате: False = стрелка списка видна
        showErrorMessage=True,
        errorTitle="Недопустимое значение",
        error="Выберите значение из списка: " + ", ".join(values),
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def link_to(cell, sheet: str, row: int) -> None:
    """Внутренняя ссылка на ячейку другого листа книги.

    Задаётся через location, а не через target: target — это внешний адрес, Excel
    открыл бы по нему файл, а не перешёл внутри книги.

    Перелинковка не украшение: без неё читатель отчёта ищет кейс по номеру глазами,
    а руками эти ссылки ставить нельзя — в первом же отчёте, размеченном вручную,
    три из них вели не туда. Ставит генератор, проверяет verify-qa-report.py.
    """
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet}'!A{row}")
    cell.font = LINK_FONT


def evidence_case_id(reference: str) -> str | None:
    """Какому кейсу принадлежит файл доказательства.

    Имя файла начинается с Case ID: TC-001.png, TC-001-2-hover.png, TC-007_hint.png.
    Путь может содержать каталог и любой разделитель после идентификатора.
    Возвращает None для файлов, не привязанных к кейсу, — журналов и замеров.
    """
    name = str(reference).strip().replace("\\", "/").rsplit("/", 1)[-1]
    match = re.match(r"^(TC-\d{3,})(?=[^0-9]|$)", name, re.IGNORECASE)
    return match.group(1).upper() if match else None


def row_links(defect_link: str, case_id: str) -> str:
    """Ссылки на доказательства для одной строки дефекта.

    Раньше в каждую строку группы шёл один и тот же список на весь дефект: строка про
    TC-021 ссылалась на снимки TC-022, TC-023 и TC-024, и разработчик открывал не тот
    экран. Оставляем снимки этой строки плюс файлы, не привязанные к кейсу, — журнал
    прогона нужен каждой строке.
    """
    parts = [item.strip() for item in str(defect_link or "").split(",") if item.strip()]
    if not parts:
        return ""

    kept = [p for p in parts if evidence_case_id(p) in (None, case_id.upper())]
    # Если по кейсу не нашлось ничего, отдаём список целиком: у старых входных файлов
    # имена доказательств могли не следовать правилу, и терять их нельзя.
    return ", ".join(kept or parts)


def case_row_map(cases: list) -> dict:
    """Case ID → строка на листе Test Cases. Порядок гарантирован validate_input."""
    return {str(case["id"]).strip(): index for index, case in enumerate(cases, start=2)}


def defect_row_map(defects: list) -> dict:
    """Bug ID → строка объявления дефекта на листе Bug Reports.

    Дефект занимает столько строк, сколько у него rows, а идентификатор несёт только
    первая. Карта считается заранее: Test Run строится раньше Bug Reports и должен
    знать, куда будет вести ссылка.
    """
    rows = {}
    cursor = 2
    for defect in defects:
        rows[str(defect["id"]).strip()] = cursor
        cursor += len(defect.get("rows") or [])
    return rows


def add_table(ws, name: str, last_row: int, col_count: int) -> None:
    ref = f"A1:{get_column_letter(col_count)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table)


def build_data_sheet(wb) -> None:
    """Скрытый лист справочников. Именованные диапазоны переносимее, чем
    ссылка вида =Data!$B$1:$B$4 — старые Excel их в валидации не принимают."""
    ws = wb.create_sheet("Data")
    row = 1
    for label, values in LOOKUPS:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        first = row
        for value in values:
            ws.cell(row=row, column=2, value=value).font = Font(size=11)
            row += 1
        last = row - 1
        wb.defined_names.add(DefinedName(label, attr_text=f"Data!$B${first}:$B${last}"))
        row += 1  # пустая строка-разделитель между справочниками
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.sheet_state = "hidden"


def build_cases_sheet(wb, cases: list) -> None:
    ws = wb.create_sheet("Test Cases")
    style_header(ws, CASE_HEADERS)

    for index, case in enumerate(cases, start=2):
        ws.cell(row=index, column=1, value=str(case["id"]).strip())
        ws.cell(row=index, column=2, value=require(case, "title", f"Кейс {case['id']}"))
        ws.cell(row=index, column=3, value=require(case, "preconditions", f"Кейс {case['id']}"))
        ws.cell(row=index, column=4, value=join_steps(case.get("steps"), f"steps кейса {case['id']}"))
        ws.cell(row=index, column=5, value=require(case, "expected", f"Кейс {case['id']}"))
        ws.cell(row=index, column=6, value=str(case.get("postconditions") or "").strip())
        ws.cell(row=index, column=7, value=str(case["priority"]).strip())

    last_row = len(cases) + 1
    style_body(ws, 2, last_row, len(CASE_HEADERS), CENTERED_CASE_COLS)
    fit_rows(ws, 2, last_row, fit_columns(ws, CASE_HEADERS, 2, last_row))
    add_table(ws, "TestCasesTable", last_row, len(CASE_HEADERS))

    add_dropdown(ws, "PriorityOptions", PRIORITY_VALUES, f"G2:G{last_row}")


def build_run_sheet(wb, data: dict, run: list, case_rows: dict, defect_rows: dict) -> None:
    ws = wb.create_sheet("Test Run")
    style_header(ws, RUN_HEADERS)

    session = data.get("session") or {}
    session_id = require(session, "id", "Секция session")
    session_name = require(session, "name", "Секция session")
    tester = str(data.get("tester") or "").strip() or DEFAULT_TESTER
    session_date = parse_date(session.get("date"), "session.date")

    for index, row in enumerate(run, start=2):
        # Session ID проставляется только в первой строке: в эталонном отчёте
        # сессия читается как один блок, а не повторяется в каждой строке.
        if index == 2:
            ws.cell(row=index, column=1, value=session_id)
        date_cell = ws.cell(row=index, column=2, value=session_date)
        date_cell.number_format = DATE_FORMAT
        ws.cell(row=index, column=3, value=session_name)
        ws.cell(row=index, column=4, value=tester)

        case_id = str(row["case"]).strip()
        case_cell = ws.cell(row=index, column=5, value=case_id)
        if case_id in case_rows:
            link_to(case_cell, "Test Cases", case_rows[case_id])

        ws.cell(row=index, column=6, value=str(row["result"]).strip())
        ws.cell(row=index, column=7, value=str(row.get("comment") or "").strip())

        bug = str(row.get("bug") or "").strip()
        if bug:
            bug_cell = ws.cell(row=index, column=8, value=bug)
            if bug in defect_rows:
                link_to(bug_cell, "Bug Reports", defect_rows[bug])

    last_row = len(run) + 1
    style_body(ws, 2, last_row, len(RUN_HEADERS), CENTERED_RUN_COLS)
    fit_rows(ws, 2, last_row, fit_columns(ws, RUN_HEADERS, 2, last_row))
    add_table(ws, "TestRunTable", last_row, len(RUN_HEADERS))

    add_dropdown(ws, "ResultOptions", RESULT_VALUES, f"F2:F{last_row}")

    for value, (fill, font_color) in RESULT_STYLES.items():
        ws.conditional_formatting.add(
            f"F2:F{last_row}",
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill("solid", bgColor=fill),
                font=Font(color=font_color, bold=True, size=10),
            ),
        )


def build_bugs_sheet(wb, data: dict, defects: list, case_rows: dict) -> None:
    ws = wb.create_sheet("Bug Reports")
    style_header(ws, BUG_HEADERS)

    session_id = require(data.get("session") or {}, "id", "Секция session")
    default_env = require(data, "environment", "Корень отчёта")

    row_index = 2
    for defect in defects:
        defect_id = str(defect["id"]).strip()
        title = require(defect, "title", f"Дефект {defect_id}")
        environment = str(defect.get("environment") or default_env).strip()
        link = str(defect.get("link") or "").strip()
        reported = parse_date(defect.get("reported") or (data.get("session") or {}).get("date"), f"reported дефекта {defect_id}")

        for position, item in enumerate(defect["rows"]):
            # Идентификатор и название несёт только первая строка группы —
            # так один дефект читается как единый блок, разложенный по элементам.
            if position == 0:
                ws.cell(row=row_index, column=1, value=defect_id)
                ws.cell(row=row_index, column=2, value=title)
            related = str(item.get("case") or "").strip()
            related_cell = ws.cell(row=row_index, column=3, value=related)
            if related in case_rows:
                link_to(related_cell, "Test Cases", case_rows[related])

            # Session ID на листе Test Run стоит только в первой строке прогона —
            # туда и ведём, иначе ссылка упиралась бы в пустую ячейку.
            session_cell = ws.cell(row=row_index, column=4, value=session_id)
            link_to(session_cell, "Test Run", 2)
            ws.cell(row=row_index, column=5, value=require(item, "preconditions", f"Дефект {defect_id}"))
            ws.cell(row=row_index, column=6, value=require(item, "step", f"Дефект {defect_id}"))
            ws.cell(row=row_index, column=7, value=require(item, "expected", f"Дефект {defect_id}"))
            ws.cell(row=row_index, column=8, value=require(item, "actual", f"Дефект {defect_id}"))
            ws.cell(row=row_index, column=9, value=str(defect["severity"]).strip())
            ws.cell(row=row_index, column=10, value=str(defect["priority"]).strip())
            ws.cell(row=row_index, column=11, value=environment)

            # Своя ссылка строки имеет приоритет; иначе список дефекта фильтруется по кейсу.
            row_link = str(item.get("link") or "").strip() or row_links(link, related)
            link_cell = ws.cell(row=row_index, column=12, value=row_link)
            if row_link.startswith(("http://", "https://")):
                link_cell.hyperlink = row_link
                link_cell.font = LINK_FONT

            date_cell = ws.cell(row=row_index, column=13, value=reported)
            date_cell.number_format = DATE_FORMAT
            row_index += 1

    last_row = row_index - 1
    if last_row < 2:
        # Книга без дефектов легальна: прогон мог пройти полностью зелёным.
        last_row = 2
    else:
        style_body(ws, 2, last_row, len(BUG_HEADERS), CENTERED_BUG_COLS)
        fit_rows(ws, 2, last_row, fit_columns(ws, BUG_HEADERS, 2, last_row))
        add_table(ws, "BugReportsTable", last_row, len(BUG_HEADERS))

        add_dropdown(ws, "SeverityOptions", SEVERITY_VALUES, f"I2:I{last_row}")
        add_dropdown(ws, "PriorityOptions", PRIORITY_VALUES, f"J2:J{last_row}")


def build_workbook(data: dict):
    validate_input(data)

    wb = Workbook()
    wb.remove(wb.active)

    defects = data.get("defects") or []
    # Карты строк считаются до сборки: Test Run строится раньше Bug Reports,
    # а ссылаться должен на строки, которых ещё нет.
    case_rows = case_row_map(data["cases"])
    defect_rows = defect_row_map(defects)

    build_data_sheet(wb)
    build_cases_sheet(wb, data["cases"])
    build_run_sheet(wb, data, data["run"], case_rows, defect_rows)
    build_bugs_sheet(wb, data, defects, case_rows)

    return wb


NUMERIC_XML_ENTITY = re.compile(r"&#(?:x([0-9A-Fa-f]+)|([0-9]+));")


def normalize_xlsx_unicode(path: Path) -> int:
    """Заменить не-ASCII XML-коды в XLSX на обычный UTF-8.

    Некоторые сборки Excel под Windows показывают числовые XML-сущности из inline-строк
    буквально. OOXML допускает обычный UTF-8, поэтому перепаковываем только XML-части книги,
    сохраняя метаданные ZIP-записей. Исходный файл заменяется лишь после проверки архива.
    """
    replacements = 0

    def decode_non_ascii_entity(match: re.Match[str]) -> str:
        nonlocal replacements
        codepoint = int(match.group(1), 16) if match.group(1) else int(match.group(2), 10)
        if codepoint < 128 or codepoint > 0x10FFFF or 0xD800 <= codepoint <= 0xDFFF:
            return match.group(0)
        replacements += 1
        return chr(codepoint)

    with tempfile.NamedTemporaryFile(
        prefix=f".{path.stem}-", suffix=".xlsx.tmp", dir=path.parent, delete=False
    ) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        with zipfile.ZipFile(path, "r") as source_zip, zipfile.ZipFile(temp_path, "w") as target_zip:
            for info in source_zip.infolist():
                content = source_zip.read(info.filename)
                if info.filename.endswith(".xml"):
                    xml = content.decode("utf-8")
                    content = NUMERIC_XML_ENTITY.sub(decode_non_ascii_entity, xml).encode("utf-8")
                target_zip.writestr(info, content)

        with zipfile.ZipFile(temp_path, "r") as check_zip:
            damaged_entry = check_zip.testzip()
            if damaged_entry:
                raise RuntimeError(f"Повреждённый элемент XLSX: {damaged_entry}")

        temp_path.replace(path)
    finally:
        temp_path.unlink(missing_ok=True)

    return replacements


def save_workbook(wb, path: Path) -> int:
    """Сохранить книгу в совместимом с Windows Excel OOXML."""
    wb.save(path)
    return normalize_xlsx_unicode(path)


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Собрать XLSX-отчёт Test Management из JSON.")
    parser.add_argument("--input", required=True, help="JSON с кейсами, прогоном и дефектами.")
    parser.add_argument("--out", required=True, help="Путь к выходному .xlsx")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Не найден входной файл: {input_path}", file=sys.stderr)
        return 2

    try:
        data = json.loads(input_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"Входной файл не разбирается как JSON: {exc}", file=sys.stderr)
        return 2

    try:
        wb = build_workbook(data)
    except ReportError as exc:
        print(f"Отчёт не собран: {exc}", file=sys.stderr)
        return 1

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    converted_entities = save_workbook(wb, out_path)

    cases = len(data["cases"])
    passed = sum(1 for row in data["run"] if row["result"] == "Pass")
    failed = sum(1 for row in data["run"] if row["result"] == "Fail")
    print(f"Собран {out_path}")
    print(f"Кейсов: {cases} | Pass: {passed} | Fail: {failed} | Дефектов: {len(data.get('defects') or [])}")
    print(f"Совместимость Windows Excel: преобразовано XML-кодов: {converted_entities}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
