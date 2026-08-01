"""
verify-qa-report — независимая проверка собранного XLSX-отчёта.

Читает готовый файл, а не исходный JSON: иначе проверка повторяла бы логику
генератора и не поймала бы его собственную ошибку. Инварианты перенесены из
валидатора, написанного под первый реальный отчёт этого формата.

Использование:
    python tools/verify-qa-report.py QA_Проект_2026-07-30.xlsx
    python tools/verify-qa-report.py отчёт.xlsx --expect-pass 10 --expect-fail 5
    python tools/verify-qa-report.py старый-отчёт.xlsx --legacy
    python tools/verify-qa-report.py отчёт.xlsx --evidence <проект>/evidence
    python tools/verify-qa-report.py отчёт.xlsx --summary <проект>/outputs/ОТЧЁТ.md
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Не установлен openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

EXPECTED_SHEETS = ["Data", "Test Cases", "Test Run", "Bug Reports"]

EXPECTED_CASE_HEADERS = [
    "Case ID\n(Номер кейса)",
    "Title\n(Название кейса)",
    "Preconditions\n(Предусловия)",
    "Steps\n(Шаги)",
    "Expected Result\n(Ожидаемый результат)",
    "Postconditions\n(Постусловия)",
    "Priority\n(Приоритет)",
]

EXPECTED_RUN_HEADERS = [
    "Session ID\n(Номер сессии)",
    "Date\n(Время)",
    "Session Name\n(Имя сессии)",
    "Tester\n(Тестировщик)",
    "Case ID\n(Номер кейса)",
    "Result\n(Результат)",
    "Comment\n(Комментарий)",
    "Bug ID\n(Номер бага)",
]

EXPECTED_BUG_HEADERS = [
    "Bug ID\n(Номер бага)",
    "Title\n(Название бага)",
    "Related Case ID\n(Связанные кейсы)",
    "Session ID\n(Номер сессии)",
    "Preconditions\n(Предусловия)",
    "Error Step\n(Шаг с ошибкой)",
    "Expected Result\n(Ожидаемый результат)",
    "Actual Result\n(Фактический результат)",
    "Severity\n(Серьёзность)",
    "Priority\n(Приоритет)",
    "Environment\n(Окружение)",
    "Attachments/Links\n(Скриншоты/Ссылки)",
    "Date Reported\n(Дата обнаружения)",
]

FORMULA_ERROR_RE = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A|#NULL!|#NUM!")

RESULT_VALUES = {"Pass", "Fail", "Blocked", "N/A"}
SEVERITY_VALUES = {"Critical", "Major", "Minor", "Trivial"}
PRIORITY_VALUES = {"High", "Medium", "Low"}


def text(value) -> str:
    return str(value).strip() if value is not None else ""


def column(ws, col: int, first_row: int = 2) -> list[str]:
    return [text(ws.cell(row=r, column=col).value) for r in range(first_row, ws.max_row + 1)]


def check_headers(ws, expected: list[str], failures: list[str]) -> None:
    for index, want in enumerate(expected, start=1):
        got = text(ws.cell(row=1, column=index).value)
        if got != want:
            failures.append(
                f"Лист «{ws.title}», колонка {index}: шапка «{got}», ожидается «{want}»."
            )


def link_target(cell) -> str:
    """Куда ведёт ячейка. Внутренняя ссылка живёт в location, внешняя — в target."""
    if cell.hyperlink is None:
        return ""
    return text(cell.hyperlink.location or cell.hyperlink.target)


def check_link(ws, row: int, col: int, want: str, failures: list[str]) -> None:
    cell = ws.cell(row=row, column=col)
    if not text(cell.value):
        return
    got = link_target(cell)
    if not got:
        failures.append(
            f"Нет перелинковки в «{ws.title}»!{cell.coordinate} («{text(cell.value)}»): "
            f"ожидается ссылка на {want}."
        )
    elif got != want:
        failures.append(
            f"Ссылка в «{ws.title}»!{cell.coordinate} («{text(cell.value)}») ведёт на {got}, "
            f"ожидается {want}."
        )


def check_links(run_ws, cases_ws, bugs_ws, failures: list[str]) -> None:
    """Перелинковка идентификаторов между листами.

    Ссылки ставит генератор, поэтому проверка тут не формальность: в первом отчёте,
    размеченном вручную, три ссылки из четырнадцати вели не на тот лист, а одной
    не было вовсе. Такую ошибку глазами в книге не видно — она видна только по клику.
    """
    case_rows = {
        text(cases_ws.cell(row=r, column=1).value): r
        for r in range(2, cases_ws.max_row + 1)
        if text(cases_ws.cell(row=r, column=1).value)
    }
    bug_rows = {
        text(bugs_ws.cell(row=r, column=1).value): r
        for r in range(2, bugs_ws.max_row + 1)
        if text(bugs_ws.cell(row=r, column=1).value)
    }

    for row in range(2, run_ws.max_row + 1):
        case_id = text(run_ws.cell(row=row, column=5).value)
        if case_id in case_rows:
            check_link(run_ws, row, 5, f"'Test Cases'!A{case_rows[case_id]}", failures)
        bug_id = text(run_ws.cell(row=row, column=8).value)
        if bug_id in bug_rows:
            check_link(run_ws, row, 8, f"'Bug Reports'!A{bug_rows[bug_id]}", failures)

    for row in range(2, bugs_ws.max_row + 1):
        related = text(bugs_ws.cell(row=row, column=3).value)
        if related in case_rows:
            check_link(bugs_ws, row, 3, f"'Test Cases'!A{case_rows[related]}", failures)
        # Session ID ведёт в первую строку прогона: только там он и проставлен.
        check_link(bugs_ws, row, 4, "'Test Run'!A2", failures)


EVIDENCE_CASE_RE = re.compile(r"^(TC-\d{3,})(?=[^0-9]|$)", re.IGNORECASE)


def evidence_case_id(reference: str) -> str | None:
    """Какому кейсу принадлежит файл доказательства по его имени."""
    name = text(reference).replace("\\", "/").rsplit("/", 1)[-1]
    match = EVIDENCE_CASE_RE.match(name)
    return match.group(1).upper() if match else None


def check_evidence_naming(bugs_ws, failures: list[str]) -> None:
    """В строке дефекта стоят снимки её собственного кейса.

    Один список доказательств на весь дефект приводит к тому, что строка про TC-021
    ссылается на кадры TC-022…TC-024: разработчик открывает не тот экран и не видит
    описанного. Файлы без Case ID в имени — журналы и замеры — общие для всех строк.
    """
    for row in range(2, bugs_ws.max_row + 1):
        related = text(bugs_ws.cell(row=row, column=3).value).upper()
        if not related:
            continue
        cell = bugs_ws.cell(row=row, column=12)
        for part in [p.strip() for p in text(cell.value).split(",") if p.strip()]:
            owner = evidence_case_id(part)
            if owner and owner != related:
                failures.append(
                    f"Доказательство не того кейса в «Bug Reports»!{cell.coordinate}: "
                    f"«{part}» относится к {owner}, а строка описывает {related}."
                )


def check_evidence_files(evidence_dir: Path, case_ids: list[str], failures: list[str]) -> None:
    """Каждому кейсу — свой файл, каждому файлу — свой кейс.

    Проверяется каталог, а не книга: имя, записанное в отчёт, может быть верным, а файла
    рядом не оказаться. Обход рекурсивный — часть кадров лежит в подкаталогах.
    """
    if not evidence_dir.is_dir():
        failures.append(f"Каталог доказательств не найден: {evidence_dir}")
        return

    by_case: dict[str, list[str]] = {}
    for path in sorted(evidence_dir.rglob("*")):
        if not path.is_file():
            continue
        owner = evidence_case_id(path.name)
        if owner:
            by_case.setdefault(owner, []).append(path.name)

    known = {case_id.upper() for case_id in case_ids}

    for case_id in case_ids:
        if case_id.upper() not in by_case:
            failures.append(
                f"У кейса {case_id} нет доказательства: в {evidence_dir} нет файла с именем, "
                f"начинающимся на «{case_id}»."
            )

    for owner, names in sorted(by_case.items()):
        if owner not in known:
            failures.append(
                f"Доказательство не соответствует ни одному кейсу: {', '.join(names)} — "
                f"имя указывает на {owner}, такого кейса в отчёте нет."
            )


def check_summary(
    summary_path: Path,
    case_ids: list[str],
    bug_ids: set[str],
    passed: int,
    failed: int,
    failures: list[str],
) -> None:
    """Сопроводительный текст сходится с книгой.

    Числа в тексте и в файле разошлись на реальном заказе и продержались три круга
    обсуждения: в тексте стояло 8 Pass, в книге 10. Заказчик открывает оба документа,
    и расхождение подрывает доверие ко всему отчёту — раньше, чем он дойдёт до сути.
    """
    if not summary_path.is_file():
        failures.append(f"Сопроводительный отчёт не найден: {summary_path}")
        return

    body = summary_path.read_text(encoding="utf-8")

    for label, expected in (("Pass", passed), ("Fail", failed)):
        found = [int(n) for n in re.findall(rf"(\d+)\s*{label}\b", body)]
        if not found:
            failures.append(
                f"В сопроводительном отчёте нет итога по {label}. "
                f"Ожидается строка вида «Итог: {passed} Pass, {failed} Fail.»"
            )
        elif expected not in found:
            failures.append(
                f"Итог по {label} не сходится: в тексте {', '.join(map(str, found))}, "
                f"в книге {expected}."
            )

    known_cases = {case_id.upper() for case_id in case_ids}
    for mentioned in sorted(set(m.upper() for m in re.findall(r"\bTC-\d{3,}\b", body))):
        if mentioned not in known_cases:
            failures.append(
                f"Сопроводительный отчёт ссылается на кейс {mentioned}, которого нет в книге."
            )

    known_bugs = {bug_id.upper() for bug_id in bug_ids}
    for mentioned in sorted(set(m.upper() for m in re.findall(r"\bBR-\d{3,}\b", body))):
        if mentioned not in known_bugs:
            failures.append(
                f"Сопроводительный отчёт ссылается на дефект {mentioned}, которого нет в книге."
            )


def check_tester(run_ws, failures: list[str]) -> None:
    """Отчёт подписан человеком: пустое поле Tester — отчёт без автора."""
    for row in range(2, run_ws.max_row + 1):
        if not text(run_ws.cell(row=row, column=5).value):
            continue
        if not text(run_ws.cell(row=row, column=4).value):
            failures.append(f"Строка прогона {row}: не заполнено поле Tester.")


def check_formula_errors(wb, failures: list[str]) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if FORMULA_ERROR_RE.search(str(cell.value)):
                    failures.append(f"Ошибка формулы в «{ws.title}»!{cell.coordinate}: {cell.value}")


def verify(
    path: Path,
    expect_pass: int | None,
    expect_fail: int | None,
    legacy: bool = False,
    evidence: Path | None = None,
    summary: Path | None = None,
) -> list[str]:
    failures: list[str] = []
    wb = openpyxl.load_workbook(path)

    if wb.sheetnames != EXPECTED_SHEETS:
        failures.append(
            f"Неверный состав или порядок листов: {wb.sheetnames}, ожидается {EXPECTED_SHEETS}."
        )
        return failures

    if wb["Data"].sheet_state != "hidden":
        failures.append("Лист «Data» должен быть скрытым — он служебный справочник.")

    cases_ws = wb["Test Cases"]
    run_ws = wb["Test Run"]
    bugs_ws = wb["Bug Reports"]

    check_headers(cases_ws, EXPECTED_CASE_HEADERS, failures)
    check_headers(run_ws, EXPECTED_RUN_HEADERS, failures)
    check_headers(bugs_ws, EXPECTED_BUG_HEADERS, failures)

    # --- тест-кейсы: сплошная нумерация с TC-001
    case_ids = [value for value in column(cases_ws, 1) if value]
    if not case_ids:
        failures.append("На листе «Test Cases» нет ни одного кейса.")
        return failures

    for index, case_id in enumerate(case_ids, start=1):
        expected_id = f"TC-{index:03d}"
        if case_id != expected_id:
            failures.append(
                f"Нарушена сплошная нумерация кейсов: строка {index + 1} содержит «{case_id}», ожидается «{expected_id}»."
            )

    for index in range(2, len(case_ids) + 2):
        priority = text(cases_ws.cell(row=index, column=7).value)
        if priority not in PRIORITY_VALUES:
            failures.append(f"Кейс в строке {index}: недопустимый Priority «{priority}».")
        expected_result = text(cases_ws.cell(row=index, column=5).value)
        if not expected_result:
            failures.append(f"Кейс в строке {index}: пустой Expected Result — проверить нечего.")

    # --- прогон: покрывает все кейсы в том же порядке
    run_case_ids = [value for value in column(run_ws, 5) if value]
    if run_case_ids != case_ids:
        failures.append(
            f"Test Run не покрывает все кейсы по порядку: в прогоне {len(run_case_ids)} строк, "
            f"в кейсах {len(case_ids)}; первое расхождение — "
            f"{next((f'ожидался {a}, найден {b}' for a, b in zip(case_ids, run_case_ids) if a != b), 'состав совпал, различается длина')}."
        )

    # --- дефекты: собираем объявленные Bug ID
    declared_bugs: list[str] = []
    for index in range(2, bugs_ws.max_row + 1):
        bug_id = text(bugs_ws.cell(row=index, column=1).value)
        if bug_id:
            declared_bugs.append(bug_id)
        severity = text(bugs_ws.cell(row=index, column=9).value)
        priority = text(bugs_ws.cell(row=index, column=10).value)
        actual = text(bugs_ws.cell(row=index, column=8).value)
        if not severity and not priority and not actual:
            continue  # пустая строка-хвост
        if severity not in SEVERITY_VALUES:
            failures.append(f"Дефект в строке {index}: недопустимый Severity «{severity}».")
        if priority not in PRIORITY_VALUES:
            failures.append(f"Дефект в строке {index}: недопустимый Priority «{priority}».")
        if not actual:
            failures.append(f"Дефект в строке {index}: пустой Actual Result.")

    bug_set = set(declared_bugs)
    if len(bug_set) != len(declared_bugs):
        failures.append("На листе «Bug Reports» встречаются повторяющиеся Bug ID.")

    # --- трассируемость результата и дефекта
    passed = 0
    failed = 0
    referenced_bugs: set[str] = set()
    for index in range(2, run_ws.max_row + 1):
        case_id = text(run_ws.cell(row=index, column=5).value)
        if not case_id:
            continue
        result = text(run_ws.cell(row=index, column=6).value)
        bug = text(run_ws.cell(row=index, column=8).value)

        if result not in RESULT_VALUES:
            failures.append(f"Строка прогона {case_id}: недопустимый Result «{result}».")
            continue

        if result == "Pass":
            passed += 1
            if bug:
                failures.append(f"Pass ошибочно связан с багом: {case_id} -> «{bug}».")
        elif result == "Fail":
            failed += 1
            if not bug:
                failures.append(f"Fail без корректного Bug ID: {case_id}.")
            elif bug not in bug_set:
                failures.append(f"Fail без корректного Bug ID: {case_id} ссылается на несуществующий «{bug}».")
            else:
                referenced_bugs.add(bug)

    orphan_bugs = bug_set - referenced_bugs
    if orphan_bugs:
        failures.append(
            f"Дефекты не связаны ни с одним Fail: {', '.join(sorted(orphan_bugs))}."
        )

    if expect_pass is not None and passed != expect_pass:
        failures.append(f"Итог Pass={passed}, ожидается {expect_pass}.")
    if expect_fail is not None and failed != expect_fail:
        failures.append(f"Итог Fail={failed}, ожидается {expect_fail}.")

    # Перелинковка и подпись введены позже первых заказов. Сданные книги переделывать
    # не будем, но проверять их иногда нужно — для этого и ключ. Все остальные
    # инварианты в legacy-режиме проверяются как обычно.
    if not legacy:
        check_links(run_ws, cases_ws, bugs_ws, failures)
        check_tester(run_ws, failures)
        check_evidence_naming(bugs_ws, failures)
        if evidence is not None:
            check_evidence_files(evidence, case_ids, failures)

    # Сходимость чисел проверяется и в legacy: правило действует всегда, оно не про формат.
    if summary is not None:
        check_summary(summary, case_ids, bug_set, passed, failed, failures)

    check_formula_errors(wb, failures)
    return failures


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Проверить собранный XLSX-отчёт Test Management.")
    parser.add_argument("report", help="Путь к .xlsx")
    parser.add_argument("--expect-pass", type=int, default=None)
    parser.add_argument("--expect-fail", type=int, default=None)
    parser.add_argument(
        "--legacy",
        action="store_true",
        help="Не требовать перелинковки, поля Tester и правил именования доказательств. "
             "Только для книг, собранных до введения этих правил: новые отчёты обязаны их иметь.",
    )
    parser.add_argument(
        "--evidence",
        default=None,
        help="Каталог доказательств. Сверяет файлы с кейсами: у каждого кейса есть снимок, "
             "имя которого начинается с его Case ID, и наоборот.",
    )
    parser.add_argument(
        "--summary",
        default=None,
        help="Сопроводительный отчёт (outputs/ОТЧЁТ.md). Сверяет числа Pass и Fail с книгой и "
             "проверяет, что упомянутые TC-NNN и BR-NNN в книге существуют.",
    )
    args = parser.parse_args(argv)

    path = Path(args.report)
    if not path.exists():
        print(f"Не найден файл отчёта: {path}", file=sys.stderr)
        return 2

    failures = verify(
        path,
        args.expect_pass,
        args.expect_fail,
        legacy=args.legacy,
        evidence=Path(args.evidence) if args.evidence else None,
        summary=Path(args.summary) if args.summary else None,
    )

    print(f"Проверка отчёта: {path}" + (" (режим legacy)" if args.legacy else ""))
    if failures:
        print(f"Нарушений: {len(failures)}")
        for failure in failures:
            print(f"- {failure}")
        return 1

    print("Нарушений: 0")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
