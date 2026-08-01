"""
test-qa-report — самотест связки «генератор отчёта + валидатор».

Половина тестов здесь негативные и это принципиально: валидатор, который никогда
не падает, не умеет и подтвердить успех. Поломки вносятся прямо в собранный XLSX,
а не во входной JSON — так проверяется, что валидатор действительно независим от
генератора, а не повторяет его логику.

Использование:
    python tools/test-qa-report.py
"""
from __future__ import annotations

import importlib.util
import json
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

TOOLS = Path(__file__).resolve().parent


def load(module_name: str, filename: str):
    """Имена файлов содержат дефис, обычный import их не берёт."""
    spec = importlib.util.spec_from_file_location(module_name, TOOLS / filename)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


builder = load("qa_report_builder", "build-qa-report.py")
verifier = load("qa_report_verifier", "verify-qa-report.py")


def fixture() -> dict:
    """Пять кейсов, три Pass и два Fail, один дефект на две строки."""
    cases = []
    for index in range(1, 6):
        cases.append({
            "id": f"TC-{index:03d}",
            "title": f"Шапка: проверка элемента {index}",
            "preconditions": "Открыта страница https://example.test/. Пользователь не авторизован.",
            "steps": [f"1. Навести курсор на элемент {index}.", "2. Наблюдать состояние не менее 0,3 секунды."],
            "expected": "Элемент получает заметную визуальную реакцию на наведение.",
            "postconditions": "Страница не изменена.",
            "priority": "Low" if index % 2 else "Medium",
        })

    run = []
    for index in range(1, 6):
        is_fail = index in (2, 4)
        run.append({
            "case": f"TC-{index:03d}",
            "result": "Fail" if is_fail else "Pass",
            "comment": "Стили до и после наведения совпали." if is_fail else "Реакция на наведение присутствует.",
            "bug": "BR-001" if is_fail else "",
        })

    defects = [{
        "id": "BR-001",
        "title": "Нет hover-эффекта у элементов шапки",
        "severity": "Minor",
        "priority": "Medium",
        # Список на весь дефект: генератор обязан разложить его по строкам — каждой
        # строке её снимок, а общая ссылка достаётся обеим.
        "link": "https://example.test/evidence/br-001, TC-002.png, TC-004.png",
        "reported": "2026-07-30",
        "rows": [
            {
                "case": "TC-002",
                "preconditions": "Открыта страница https://example.test/.",
                "step": "Навести курсор на элемент 2.",
                "expected": "Элемент получает заметную визуальную реакцию на наведение.",
                "actual": "Цвет, фон и рамка не изменяются.",
            },
            {
                "case": "TC-004",
                "preconditions": "Открыта страница https://example.test/.",
                "step": "Навести курсор на элемент 4.",
                "expected": "Элемент получает заметную визуальную реакцию на наведение.",
                "actual": "Цвет, фон и рамка не изменяются.",
            },
        ],
    }]

    return {
        "project": "Example",
        "environment": "Chrome 150, Windows 10",
        "tester": "Тестировщик",
        "session": {"id": "TR-001", "name": "Проверка шапки example.test", "date": "2026-07-30"},
        "cases": cases,
        "run": run,
        "defects": defects,
    }


class Results:
    def __init__(self) -> None:
        self.passed = 0
        self.failed: list[str] = []

    def ok(self, name: str) -> None:
        self.passed += 1
        print(f"  ok   {name}")

    def bad(self, name: str, detail: str) -> None:
        self.failed.append(f"{name}: {detail}")
        print(f"  FAIL {name} — {detail}")


def expect_clean(results: Results, name: str, path: Path, **kwargs) -> None:
    failures = verifier.verify(path, kwargs.get("expect_pass"), kwargs.get("expect_fail"))
    if failures:
        results.bad(name, f"ожидался чистый отчёт, получено: {failures}")
    else:
        results.ok(name)


def expect_failure(results: Results, name: str, path: Path, fragment: str) -> None:
    failures = verifier.verify(path, None, None)
    if not failures:
        results.bad(name, "валидатор не заметил поломку")
        return
    if not any(fragment.lower() in failure.lower() for failure in failures):
        results.bad(name, f"нет ожидаемой ошибки «{fragment}», получено: {failures}")
        return
    results.ok(name)


def main() -> int:
    results = Results()
    workdir = Path(tempfile.mkdtemp(prefix="qa-report-selftest-"))

    try:
        data = fixture()
        source = workdir / "report.xlsx"
        builder.build_workbook(data).save(source)

        print("Позитивные проверки")
        expect_clean(results, "чистый отчёт проходит валидатор", source, expect_pass=3, expect_fail=2)

        wb = openpyxl.load_workbook(source)
        if wb.sheetnames == ["Data", "Test Cases", "Test Run", "Bug Reports"]:
            results.ok("состав и порядок листов соответствуют спецификации")
        else:
            results.bad("состав листов", str(wb.sheetnames))
        if wb["Data"].sheet_state == "hidden":
            results.ok("лист Data скрыт")
        else:
            results.bad("лист Data", "не скрыт")
        if wb["Test Run"].freeze_panes == "A2":
            results.ok("шапка закреплена")
        else:
            results.bad("freeze panes", str(wb["Test Run"].freeze_panes))
        link = wb["Bug Reports"].cell(row=2, column=12)
        if link.hyperlink is not None:
            results.ok("ссылка на доказательство кликабельна")
        else:
            results.bad("гиперссылка", "не проставлена")

        # Перелинковка идентификаторов. Фикстура: кейсы TC-001…TC-005 в строках 2–6,
        # провалы на TC-002 и TC-004 (строки прогона 3 и 5), дефект BR-001 объявлен
        # в строке 2 листа Bug Reports и разложен на TC-002 и TC-004.
        expected_links = [
            ("Test Run", 2, 5, "'Test Cases'!A2"),
            ("Test Run", 6, 5, "'Test Cases'!A6"),
            ("Test Run", 3, 8, "'Bug Reports'!A2"),
            ("Test Run", 5, 8, "'Bug Reports'!A2"),
            ("Bug Reports", 2, 3, "'Test Cases'!A3"),
            ("Bug Reports", 3, 3, "'Test Cases'!A5"),
            ("Bug Reports", 2, 4, "'Test Run'!A2"),
            ("Bug Reports", 3, 4, "'Test Run'!A2"),
        ]
        wrong = []
        for sheet, row, col, want in expected_links:
            cell = wb[sheet].cell(row=row, column=col)
            got = cell.hyperlink.location if cell.hyperlink else None
            if got != want:
                wrong.append(f"{sheet}!{cell.coordinate}: {got} вместо {want}")
        if wrong:
            results.bad("перелинковка идентификаторов", "; ".join(wrong))
        else:
            results.ok("перелинковка идентификаторов ведёт на нужные строки")

        run_link_font = wb["Test Run"].cell(row=2, column=5).font
        if run_link_font.underline == "single":
            results.ok("ссылка в прогоне читается как ссылка")
        else:
            results.bad("стиль ссылки", "оформление тела листа затёрло вид гиперссылки")

        testers = {
            wb["Test Run"].cell(row=r, column=4).value for r in range(2, 7)
        }
        if testers == {"Тестировщик"}:
            results.ok("поле Tester заполнено во всех строках прогона")
        else:
            results.bad("поле Tester", str(testers))

        # Доказательства разложены построчно: в строке про TC-002 не должно быть кадра TC-004.
        att = {row: wb["Bug Reports"].cell(row=row, column=12).value or "" for row in (2, 3)}
        if "TC-002.png" in att[2] and "TC-004.png" not in att[2] \
           and "TC-004.png" in att[3] and "TC-002.png" not in att[3]:
            results.ok("доказательства разложены по строкам дефекта")
        else:
            results.bad("раскладка доказательств", f"строка2={att[2]!r} строка3={att[3]!r}")

        if "br-001" in att[2] and "br-001" in att[3]:
            results.ok("общая ссылка дефекта досталась обеим строкам")
        else:
            results.bad("общая ссылка", f"строка2={att[2]!r} строка3={att[3]!r}")

        # Каталог доказательств: по файлу на каждый кейс фикстуры.
        ev = workdir / "evidence"
        ev.mkdir(exist_ok=True)
        for index in range(1, 6):
            (ev / f"TC-{index:03d}.png").write_bytes(b"\x89PNG")
        (ev / "run-journal.json").write_text("{}", encoding="utf-8")
        (ev / "sub").mkdir(exist_ok=True)
        (ev / "sub" / "TC-002-2-hover.png").write_bytes(b"\x89PNG")
        if verifier.verify(source, None, None, evidence=ev):
            results.bad("каталог доказательств", "полный каталог не прошёл проверку")
        else:
            results.ok("каталог доказательств: у каждого кейса свой снимок")

        no_tester = dict(fixture())
        no_tester.pop("tester")
        default_path = workdir / "default-tester.xlsx"
        builder.build_workbook(no_tester).save(default_path)
        got_default = openpyxl.load_workbook(default_path)["Test Run"].cell(row=2, column=4).value
        if got_default == builder.DEFAULT_TESTER:
            results.ok("без поля tester подставляется имя по умолчанию")
        else:
            results.bad("умолчание Tester", f"получено «{got_default}»")
        if wb["Test Cases"].tables and wb["Test Run"].tables:
            results.ok("автофильтры на месте")
        else:
            results.bad("умные таблицы", "не созданы")

        run_dv = wb["Test Run"].data_validations.dataValidation
        if run_dv and run_dv[0].formula1 == "=ResultOptions" and run_dv[0].showErrorMessage:
            results.ok("выпадающий список Result блокирует посторонний ввод")
        else:
            results.bad("валидация Result", f"{[(d.formula1, d.showErrorMessage) for d in run_dv]}")

        if set(wb.defined_names) >= {"ResultOptions", "SeverityOptions", "PriorityOptions"}:
            results.ok("именованные диапазоны справочников созданы")
        else:
            results.bad("именованные диапазоны", str(list(wb.defined_names)))

        print("\nНегативные проверки — валидатор обязан падать")

        # 1. Порядок кейсов в прогоне
        broken = workdir / "broken-order.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        ws = wb["Test Run"]
        ws.cell(row=4, column=5).value, ws.cell(row=5, column=5).value = (
            ws.cell(row=5, column=5).value, ws.cell(row=4, column=5).value
        )
        wb.save(broken)
        expect_failure(results, "переставленные кейсы в прогоне", broken, "не покрывает все кейсы по порядку")

        # 2. Fail без Bug ID
        broken = workdir / "broken-fail.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=3, column=8).value = None
        wb.save(broken)
        expect_failure(results, "Fail без Bug ID", broken, "Fail без корректного Bug ID")

        # 3. Pass с Bug ID
        broken = workdir / "broken-pass.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=2, column=8).value = "BR-001"
        wb.save(broken)
        expect_failure(results, "Pass связан с багом", broken, "Pass ошибочно связан с багом")

        # 4. Ссылка ведёт не на тот лист. Ровно эта ошибка нашлась в отчёте,
        #    размеченном вручную: Case ID указывал на служебный лист Data.
        broken = workdir / "broken-link-target.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=3, column=5).hyperlink.location = "'Data'!A3"
        wb.save(broken)
        expect_failure(results, "ссылка Case ID ведёт не на тот лист", broken, "ведёт на 'Data'!A3")

        # 5. Ссылки нет вовсе — тоже находка из ручной разметки.
        broken = workdir / "broken-link-missing.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Bug Reports"].cell(row=2, column=3).hyperlink = None
        wb.save(broken)
        expect_failure(results, "нет ссылки в Related Case ID", broken, "Нет перелинковки")

        # 6. Отчёт без подписи автора
        broken = workdir / "broken-tester.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=2, column=4).value = None
        wb.save(broken)
        expect_failure(results, "пустое поле Tester", broken, "не заполнено поле Tester")

        # 7. Доказательство чужого кейса в строке дефекта
        broken = workdir / "broken-evidence-owner.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Bug Reports"].cell(row=2, column=12).value = "TC-004.png"
        wb.save(broken)
        expect_failure(results, "снимок чужого кейса в строке дефекта", broken, "не того кейса")

        # 8. У кейса нет доказательства
        missing = ev / "TC-003.png"
        missing.rename(ev / "_отложено.png")
        found = verifier.verify(source, None, None, evidence=ev)
        if any("нет доказательства" in f for f in found):
            results.ok("кейс без снимка")
        else:
            results.bad("кейс без снимка", f"валидатор не заметил, получено: {found}")
        (ev / "_отложено.png").rename(missing)

        # 9. Снимок несуществующего кейса — опечатка в имени или лишний файл
        stray = ev / "TC-099.png"
        stray.write_bytes(b"\x89PNG")
        found = verifier.verify(source, None, None, evidence=ev)
        if any("ни одному кейсу" in f for f in found):
            results.ok("снимок несуществующего кейса")
        else:
            results.bad("лишний снимок", f"валидатор не заметил, получено: {found}")

        # 10. Тот же лишний файл под ключом legacy проходить обязан
        if verifier.verify(source, None, None, legacy=True, evidence=ev):
            results.bad("legacy и доказательства", "ключ не снял проверку именования")
        else:
            results.ok("legacy снимает и проверку доказательств")
        stray.unlink()

        # Режим legacy: книги, собранные до введения перелинковки и подписи, проверяются
        # по остальным инвариантам. Проверяем обе стороны ключа — иначе он мог бы просто
        # глушить проверку целиком.
        legacy_ok = workdir / "legacy.xlsx"
        shutil.copy(source, legacy_ok)
        wb = openpyxl.load_workbook(legacy_ok)
        for row in range(2, 7):
            wb["Test Run"].cell(row=row, column=5).hyperlink = None
            wb["Test Run"].cell(row=row, column=4).value = None
        for row in range(2, 4):
            wb["Bug Reports"].cell(row=row, column=3).hyperlink = None
            wb["Bug Reports"].cell(row=row, column=4).hyperlink = None
        wb.save(legacy_ok)
        if verifier.verify(legacy_ok, None, None, legacy=True):
            results.bad("режим legacy", "старый отчёт не прошёл, хотя ключ это разрешает")
        else:
            results.ok("режим legacy пропускает отчёт без ссылок и подписи")
        expect_failure(results, "тот же отчёт без ключа legacy", legacy_ok, "Нет перелинковки")

        legacy_broken = workdir / "legacy-broken.xlsx"
        shutil.copy(legacy_ok, legacy_broken)
        wb = openpyxl.load_workbook(legacy_broken)
        wb["Test Run"].cell(row=2, column=6).value = "Passed"
        wb.save(legacy_broken)
        if verifier.verify(legacy_broken, None, None, legacy=True):
            results.ok("режим legacy продолжает ловить остальные нарушения")
        else:
            results.bad("режим legacy", "заглушил проверку Result, а не только ссылки")

        # 4. Fail ссылается на несуществующий дефект
        broken = workdir / "broken-ref.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=3, column=8).value = "BR-999"
        wb.save(broken)
        expect_failure(results, "ссылка на несуществующий дефект", broken, "несуществующий")

        # 5. Сломанный состав листов
        broken = workdir / "broken-sheets.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Bug Reports"].title = "Defects"
        wb.save(broken)
        expect_failure(results, "переименованный лист", broken, "состав или порядок листов")

        # 6. Изменённая шапка
        broken = workdir / "broken-header.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Cases"].cell(row=1, column=4).value = "Steps"
        wb.save(broken)
        expect_failure(results, "изменённая шапка", broken, "шапка")

        # 7. Ошибка формулы в ячейке
        broken = workdir / "broken-formula.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=2, column=7).value = "#REF!"
        wb.save(broken)
        expect_failure(results, "ошибка формулы", broken, "Ошибка формулы")

        # 8. Пропущенный кейс в нумерации
        broken = workdir / "broken-numbering.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Cases"].cell(row=4, column=1).value = "TC-007"
        wb.save(broken)
        expect_failure(results, "разрыв нумерации кейсов", broken, "нумерация")

        # 9. Дефект, ни с чем не связанный
        broken = workdir / "broken-orphan.xlsx"
        shutil.copy(source, broken)
        wb = openpyxl.load_workbook(broken)
        wb["Test Run"].cell(row=3, column=6).value = "Blocked"
        wb["Test Run"].cell(row=3, column=8).value = None
        wb["Test Run"].cell(row=5, column=6).value = "Blocked"
        wb["Test Run"].cell(row=5, column=8).value = None
        wb.save(broken)
        expect_failure(results, "дефект без единого Fail", broken, "не связаны ни с одним Fail")

        print("\nПроверки входных данных генератора")
        for name, mutate, fragment in [
            ("разрыв нумерации во входном JSON", lambda d: d["cases"][2].update({"id": "TC-009"}), "нумерац"),
            ("Fail без Bug ID во входном JSON", lambda d: d["run"][1].update({"bug": ""}), "Fail без Bug ID"),
            ("Pass с Bug ID во входном JSON", lambda d: d["run"][0].update({"bug": "BR-001"}), "Pass ошибочно связан"),
            ("недопустимый Severity", lambda d: d["defects"][0].update({"severity": "Blocker"}), "Severity"),
        ]:
            payload = json.loads(json.dumps(fixture()))
            mutate(payload)
            try:
                builder.build_workbook(payload)
            except builder.ReportError as exc:
                if fragment.lower() in str(exc).lower():
                    results.ok(name)
                else:
                    results.bad(name, f"нет ожидаемого текста «{fragment}»: {exc}")
            else:
                results.bad(name, "генератор принял заведомо некорректные данные")

    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    print()
    print(f"Пройдено: {results.passed} | Провалено: {len(results.failed)}")
    if results.failed:
        for failure in results.failed:
            print(f"- {failure}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
